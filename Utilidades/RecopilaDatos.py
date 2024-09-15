# Recopila informacion del fichero "resultado_x_x_x_x.txt", y junto con nuevos calculos
# los agrega como una nueva línea al final del fichero "TablaDeResultados.xlsx"

import re
import sys
import openpyxl
import os
from openpyxl import Workbook

def agregar_valores_excel(fichero_excel, valor1, valor2, valor3, valor4, valor5, valor6, valor7, valor8, valor9, valor10,):
    try:
        # Cargar el archivo Excel existente
        libro = openpyxl.load_workbook(fichero_excel)

        # Seleccionar la hoja activa (la primera hoja en este caso)
        hoja = libro.active

        # Encontrar la última fila
        ultima_fila = hoja.max_row + 1

        # Agregar los valores a la nueva fila
        hoja.cell(row=ultima_fila, column=1, value=valor1)
        hoja.cell(row=ultima_fila, column=2, value=valor2)
        hoja.cell(row=ultima_fila, column=3, value=valor3)
        hoja.cell(row=ultima_fila, column=4, value=valor4)
        hoja.cell(row=ultima_fila, column=5, value=valor5)
        hoja.cell(row=ultima_fila, column=6, value=valor6)
        hoja.cell(row=ultima_fila, column=7, value=valor7)
        hoja.cell(row=ultima_fila, column=8, value=valor8)
        hoja.cell(row=ultima_fila, column=9, value=valor9)
        hoja.cell(row=ultima_fila, column=10, value=valor10)

        # Guardar los cambios en el archivo Excel
        libro.save(fichero_excel)
        print(f"Valores {valor1} y {valor2} agregados correctamente a {fichero_excel}.")
    
    except Exception as e:
        print(f"Error al agregar valores: {e}")

# Extrae los datos de los tiempos optimos
def extraer_valores(cadena):
    # Usar expresión regular para capturar la lista numérica en la primera línea
    patron = r"Línea con el menor valor en la última columna:\[(.+?)\]"
    coincidencia = re.search(patron, cadena)

    if coincidencia:
        # Convertir la lista en una lista de números flotantes
        lista = list(map(float, coincidencia.group(1).split(',')))
        return int(lista[-1]), int(lista[-3]), int(lista[-2])
    else:
        return None

# Extrae los datos del tiempo que ejecucion de la simulacion para la propuesta de fase con distancia euclídea
def extraer_valores_solucion_euclidea(cadena):
    # Usar expresión regular para capturar los valores en la línea de "Solución propuesta"
    patron = r"Solución propuesta\s+.+==\s+\[(\d+),\s*(\d+)\].+==>\s+([\d\.]+)"
    coincidencia = re.search(patron, cadena)

    if coincidencia:
        #return fase_euclidea_1, fase_euclidea_2, tiempo_simulacion_euclidea
        return int(coincidencia.group(1)), int(coincidencia.group(2)), int(float(coincidencia.group(3)))
    else:
        return None

# Extrae los datos del tiempo que ejecucion de la simulacion para la propuesta de fase sin distancia euclídea    
def extraer_valores_solucion_no_euclidea(cadena):
    # Usar expresión regular para capturar los valores en la línea de "Solución propuesta sin distancia euclidea"
    patron = r"Solución propuesta sin distancia euclidea\s+(\[\'[0-9\.,\' ]+\])"
    coincidencia = re.search(patron, contenido)
    if coincidencia:
        # Extraer la lista que coincide
        lista_str = coincidencia.group(1)

        # Elimina las comillas
        # Paso 1: Eliminar los corchetes y las comillas simples
        cadena_sin_corchetes = lista_str.replace("'", "").replace("[", "").replace("]", "")

        # Paso 2: Crear una lista de enteros
        lista_enteros = [int(x.strip()) for x in cadena_sin_corchetes.split(",")]

        # Paso 3: Reconstruir la cadena en el formato deseado
        cadena_nueva = str(lista_enteros).replace(" ", "")
    else:
        return None
    
    patron = r"Solución propuesta sin distancia euclidea\s+.+==\s+\[(\d+),\s*(\d+)\].+==>\s+([\d\.]+)"
    coincidencia = re.search(patron, cadena)

    if coincidencia:
        print()
        #return fase_no_euclidea_1, fase_no_euclidea_2, tiempo_simulacion_no_euclidea
        return cadena_nueva, int(coincidencia.group(1)), int(coincidencia.group(2)), int(float(coincidencia.group(3)))
    else:
        return None


def leer_fichero(fichero):
    with open(fichero, 'r', encoding='latin-1') as f:
        contenido = f.read()
    return contenido


def calcular_porcentaje_diferencia(valorCorrecto, valorIncorrecto):
    diferencia = abs(valorIncorrecto - valorCorrecto)
    porcentaje_diferencia = (diferencia / valorCorrecto) * 100
    return porcentaje_diferencia



#########################################
# MAIN
#########################################

if (len(sys.argv)==1):
    print("Es necesario argumentos")
    exit()
argumentos = sys.argv[1:]

contenido = leer_fichero(argumentos[0])

resultado = extraer_valores(contenido)
if resultado:
    valor_minimo, fase_minima_1, fase_minima_2 = resultado
    print(f"Valor mínimo: {valor_minimo}")
    print(f"Fase mínima 1: {fase_minima_1}")
    print(f"Fase mínima 2: {fase_minima_2}")
else:
    print("No se encontraron los valores.")


resultado = extraer_valores_solucion_euclidea(contenido)
if resultado:
    fase_euclidea_1, fase_euclidea_2, tiempo_simulacion_euclidea = resultado
    print(f"Fase euclidea 1: {fase_euclidea_1}")
    print(f"Fase euclidea 2: {fase_euclidea_2}")
    print(f"Tiempo de simulación euclidea: {tiempo_simulacion_euclidea}")
else:
    print("No se encontraron los valores.")

resultado = extraer_valores_solucion_no_euclidea(contenido)
if resultado:
    num_vehiculos, fase_no_euclidea_1, fase_no_euclidea_2, tiempo_simulacion_no_euclidea = resultado
    print(f"Caso: {num_vehiculos}")
    print(f"Fase no euclidea 1: {fase_no_euclidea_1}")
    print(f"Fase no euclidea 2: {fase_no_euclidea_2}")
    print(f"Tiempo de simulación no euclidea: {tiempo_simulacion_no_euclidea}")
else:
    print("No se encontraron los valores.")

porcentaje_1 = calcular_porcentaje_diferencia(valor_minimo, tiempo_simulacion_euclidea )
print(f"Porcentaje de diferencia: {porcentaje_1:.2f}%")

porcentaje_2 = calcular_porcentaje_diferencia(valor_minimo, tiempo_simulacion_no_euclidea )
print(f"Porcentaje de diferencia: {porcentaje_2:.2f}%")


# Ejemplo de uso
fichero_excel = "TablaDeResultados.xlsx"

# Comprobar si el archivo existe
if not os.path.exists(fichero_excel):
    # Si no existe, crear el archivo excel
    wb = Workbook()
    wb.save(fichero_excel)
    print(f'Archivo {fichero_excel} creado.')
else:
    print(f'El archivo {fichero_excel} ya existe.')

agregar_valores_excel(fichero_excel, num_vehiculos, valor_minimo, str([fase_minima_1, fase_minima_2]),
                       tiempo_simulacion_euclidea, str([fase_euclidea_1, fase_euclidea_2]),
                       tiempo_simulacion_no_euclidea, str([fase_no_euclidea_1, fase_no_euclidea_2]),
                       round(porcentaje_1, 2), round(porcentaje_2, 2), round(porcentaje_1-porcentaje_2,2))